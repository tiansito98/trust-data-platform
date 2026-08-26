# -*- coding: utf-8 -*-
"""Agrega al Excel la hoja 'Como leer la tabla' con el significado de cada columna."""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOCS = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'docs'))
DEST = os.path.join(DOCS, 'auditoria_comisiones_julio_REVISADO.xlsx')

INK  = "1F2937"
HDR  = PatternFill("solid", fgColor="1F3A5F")
HDRF = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
ORIG = PatternFill("solid", fgColor="DCE6F1")
CALC = PatternFill("solid", fgColor="EDEFF2")
THIN = Side(style="thin", color="BFC7D1")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# columna, origen, que es, de donde sale
DICC = [
    ("A", "ASESOR", "Original",
     "El asesor tal como aparece en el archivo que ellos entregaron.",
     "Excel de los asesores. Steffany = Jeimmy Fajardo."),
    ("B", "# CONTRATO", "Original",
     "El numero de contrato tal como ellos lo escribieron, sin corregir.",
     "Excel de los asesores."),
    ("C", "ADICIONAL", "Original",
     "La etiqueta que ellos le pusieron al adicional (BF, SL, SILLA BEBE, UP...).",
     "Excel de los asesores."),
    ("D", "VALOR", "Original",
     "El valor en pesos que ellos reclaman. OJO: es la COMISION que piden, no el precio "
     "del adicional. Por eso es mucho menor que el valor del cargo.",
     "Excel de los asesores."),
    ("E", "CONTRATO VERIFICADO", "Revision",
     "El numero de contrato que realmente existe en Sixt. Si ellos se equivocaron al "
     "escribirlo, aqui va el correcto. Si no existe ninguno parecido, dice NO EXISTE.",
     "Tabla de contratos de Sixt (bronze)."),
    ("F", "CODIGO EN SISTEMA", "Revision",
     "El codigo real del cargo en Sixt. Sirve cuando el asesor uso otro nombre: por "
     "ejemplo 'SILLA BEBE' es CS, y dos 'UP' resultaron ser FI y OT.",
     "Cargos del contrato (bronze), sin duplicados y con la ultima correccion vigente."),
    ("G", "DIAS", "Revision",
     "Cuantas unidades tiene el cargo. Casi siempre son dias de renta. Sirve para "
     "detectar cuando el asesor liquido menos dias de los que tiene el contrato.",
     "Cantidad del cargo en Sixt."),
    ("H", "VALOR CARGO USD", "Revision",
     "El valor total del adicional en dolares, sin IVA, tal como lo factura Sixt.",
     "Valor del cargo en Sixt."),
    ("I", "SE VENDIO EN COUNTER", "Revision",
     "LA COLUMNA CLAVE. Compara el cargo del contrato contra el mismo cargo en la reserva "
     "del cliente. Si ya venia en la reserva, el cliente lo compro por internet y nadie lo "
     "vendio en el counter. Valores: 'SI - walk-in' (el cliente llego sin reserva, todo es "
     "counter), 'SI - no venia en la reserva' (habia reserva pero el adicional se vendio en "
     "el mostrador), 'NO - venia incluido en la reserva' (no genera comision de counter), "
     "'PARCIAL' (una parte venia de la reserva y otra se vendio en el counter).",
     "Cargos del contrato contra cargos de la reserva, cruzados por codigo."),
    ("J", "VALOR COUNTER COP", "Revision",
     "Solo la parte del adicional que se vendio en el counter, convertida a pesos con la "
     "TRM del Banco de la Republica del dia de la entrega. Sin IVA.",
     "Valor del cargo menos lo que venia en la reserva, por la TRM Banrep."),
    ("K", "COMISION QUE CORRESPONDE", "Revision",
     "La comision que le toca al asesor: 5% sobre el valor con IVA, o sea la columna J "
     "multiplicada por 1,19 y luego por 5%. Queda en cero si el adicional no se vendio en "
     "el counter o si lo entrego otro asesor.",
     "Calculo. Es la misma regla que usan ellos y la que esta documentada en el tablero."),
    ("L", "DIFERENCIA", "Revision",
     "Columna K menos columna D. En POSITIVO le debemos mas de lo que pidio. En NEGATIVO "
     "pidio mas de lo que el sistema respalda.",
     "Calculo."),
    ("M", "QUIEN ENTREGO", "Revision",
     "Quien entrego el vehiculo de verdad, es decir quien atendio al cliente en el counter "
     "y le vendio los adicionales. Esta es la persona a la que le corresponde la comision.",
     "Tabla de vehiculos de Sixt, campo de entrega del primer tramo. Fuente sin ambiguedad."),
    ("N", "LE PAGA HOY EL SISTEMA", "Revision",
     "A quien le esta pagando la comision el tablero hoy. Cuando esta columna es distinta "
     "de la M, ahi esta el error: el sistema le paga a quien RECIBIO el carro en la "
     "devolucion, no a quien lo entrego.",
     "El campo que usa hoy el tablero para liquidar comisiones."),
    ("O", "LE CORRESPONDE", "Revision",
     "SI cuando el adicional se vendio en el counter Y lo entrego el asesor que lo reclama. "
     "NO en cualquier otro caso.",
     "Resume las columnas I y M."),
    ("P", "RESULTADO", "Revision",
     "El veredicto de la linea, en una etiqueta. Ver la tabla de abajo.",
     "Calculo."),
    ("Q", "QUE PASO", "Revision",
     "La explicacion en palabras: que encontramos, si el error fue del sistema o del asesor, "
     "y que hay que hacer.",
     "Redactado a partir de las columnas anteriores."),
]

RESULTADOS = [
    ("CORRECTO", 17,
     "El adicional existe, se vendio en el counter, lo entrego el asesor que lo reclama y "
     "el monto que pidio coincide con el del sistema. No hay nada que discutir: solo falta "
     "corregir a quien se le paga."),
    ("SUYO - FALTA DEFINIR OT/FI", 7,
     "Igual que CORRECTO, pero el cargo es de codigo OT o FI, que hoy no estan en la lista "
     "de codigos que comisionan. El asesor no se equivoco; falta que el negocio decida si "
     "esos codigos generan comision."),
    ("SUYO - REVISAR MONTO", 3,
     "El adicional es suyo y se vendio en el counter, pero el valor que reclamo no coincide "
     "con el del sistema. Dos casos son a favor del asesor (pidio de menos) y uno en contra "
     "(pidio de mas). Hay que hablar con la persona."),
    ("NO LE CORRESPONDE", 1,
     "El adicional existe y se vendio en el counter, pero lo entrego otro asesor. "
     "Caso 9523826011: lo reclaman Danilo y Natalia, y lo entrego Natalia."),
    ("NO ES COUNTER", 1,
     "El adicional venia incluido en la reserva del cliente. Nadie lo vendio en el "
     "mostrador, asi que no genera comision de counter para ningun asesor."),
    ("CONTRATO NO EXISTE", 1,
     "El numero de contrato no aparece en Sixt. Hay que pedirle el numero correcto al asesor."),
    ("FUERA DEL SISTEMA", 1,
     "No es un cargo de Sixt (el bono de resenas de Google). Se liquida por fuera."),
]


def tabla(ws, df, fila_ini, anchos, wrap):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=fila_ini, column=j, value=col)
        c.fill, c.font, c.border = HDR, HDRF, BORDE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False), start=fila_ini + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDE
            c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(wrap_text=(df.columns[j - 1] in wrap), vertical="top")
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def main():
    wb = load_workbook(DEST)
    if "Como leer la tabla" in wb.sheetnames:
        del wb["Como leer la tabla"]
    ws = wb.create_sheet("Como leer la tabla", 1)

    ws["A1"] = "COMO LEER LA HOJA 'REVISION POR LINEA'"
    ws["A1"].font = Font(bold=True, size=14, color=INK, name="Calibri")
    ws["A2"] = ("Las columnas A a D son el archivo original de los asesores, sin tocar. "
                "De la E en adelante es la revision contra el sistema Sixt.")
    ws["A2"].font = Font(size=10, italic=True, color="5A6472", name="Calibri")
    ws.merge_cells("A2:E2")

    d = pd.DataFrame(DICC, columns=["COL", "NOMBRE DE LA COLUMNA", "ORIGEN",
                                    "QUE SIGNIFICA", "DE DONDE SALE"])
    tabla(ws, d, 4, [6, 26, 11, 88, 52], {"QUE SIGNIFICA", "DE DONDE SALE"})
    for i in range(5, 5 + len(d)):
        ws.row_dimensions[i].height = 60
        ws.cell(row=i, column=3).fill = ORIG if ws.cell(row=i, column=3).value == "Original" else CALC
    ws.row_dimensions[12].height = 118          # la fila de SE VENDIO EN COUNTER

    fila = 4 + len(d) + 3
    ws.cell(row=fila - 1, column=1, value="VALORES DE LA COLUMNA 'RESULTADO'"
            ).font = Font(bold=True, size=12, color=INK, name="Calibri")
    r = pd.DataFrame(RESULTADOS, columns=["RESULTADO", "CUANTAS LINEAS", "QUE QUIERE DECIR"])
    tabla(ws, r, fila, [6, 26, 11, 88, 52], {"QUE QUIERE DECIR"})
    ws.cell(row=fila, column=1).value = "RESULTADO"
    for i in range(fila + 1, fila + 1 + len(r)):
        ws.row_dimensions[i].height = 46

    wb.save(DEST)
    print("Hojas:", wb.sheetnames)
    print("Filas diccionario:", len(d), "| valores de RESULTADO:", len(r))


if __name__ == "__main__":
    main()
