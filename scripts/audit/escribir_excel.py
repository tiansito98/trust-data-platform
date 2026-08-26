# -*- coding: utf-8 -*-
"""Escribe docs/auditoria_comisiones_julio_REVISADO.xlsx con formato."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from generar_excel import build, ASESOR_COD, nm
from claims_julio import SISTEMA_OK, TOTALES_DECLARADOS

OUT  = os.path.join(os.path.dirname(__file__), 'out')
DOCS = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'docs'))
DEST = os.path.join(DOCS, 'auditoria_comisiones_julio_REVISADO.xlsx')

ctx = pd.read_csv(os.path.join(OUT, 'x_ctx.csv')).set_index('contrato')

INK   = "1F2937"
HDR   = PatternFill("solid", fgColor="1F3A5F")
HDRF  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
ORIG  = PatternFill("solid", fgColor="DCE6F1")     # columnas del Excel original
VERDE = PatternFill("solid", fgColor="D8EFDC")
AMAR  = PatternFill("solid", fgColor="FDF0CE")
ROJO  = PatternFill("solid", fgColor="F8DCD8")
GRIS  = PatternFill("solid", fgColor="EDEFF2")
THIN  = Side(style="thin", color="BFC7D1")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLOR_RES = {
    "CORRECTO": VERDE,
    "SUYO - FALTA DEFINIR OT/FI": AMAR,
    "SUYO - REVISAR MONTO": AMAR,
    "NO ES COUNTER": ROJO,
    "NO LE CORRESPONDE": ROJO,
    "CONTRATO NO EXISTE": ROJO,
    "EL CARGO NO EXISTE": ROJO,
    "FUERA DEL SISTEMA": GRIS,
}


def escribir_tabla(ws, df, anchos, wrap_cols=(), money_cols=(), fila_ini=1):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=fila_ini, column=j, value=col)
        c.fill, c.font, c.border = HDR, HDRF, BORDE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False), start=fila_ini + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDE
            c.font = Font(size=10, name="Calibri")
            col = df.columns[j - 1]
            if col in wrap_cols:
                c.alignment = Alignment(wrap_text=True, vertical="top")
            elif col in money_cols:
                c.alignment = Alignment(horizontal="right", vertical="top")
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
            else:
                c.alignment = Alignment(vertical="top")
    for col, w in anchos.items():
        if col in list(df.columns):
            ws.column_dimensions[get_column_letter(list(df.columns).index(col) + 1)].width = w


def main():
    rev = build()
    wb = Workbook()

    # ---------- Hoja 1: revision por linea ----------
    ws = wb.active
    ws.title = "Revision por linea"
    ws["A1"] = "REVISION DE COMISIONES POR ADICIONALES - JULIO 2026"
    ws["A1"].font = Font(bold=True, size=14, color=INK, name="Calibri")
    ws["A2"] = ("Las 4 primeras columnas son el Excel original de los asesores. "
                "El resto es la revision contra el sistema Sixt, suponiendo ya corregidos "
                "los errores. Comision = 5% sobre el cargo con IVA, en pesos a TRM Banrep "
                "del dia de entrega. PENDIENTE DE APLICAR: los arreglos aun no estan en produccion.")
    ws["A2"].font = Font(size=10, italic=True, color="5A6472", name="Calibri")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:Q3")
    ws.row_dimensions[2].height = 30

    anchos = {"ASESOR": 11, "# CONTRATO": 13, "ADICIONAL": 14, "VALOR": 11,
              "CONTRATO VERIFICADO": 15, "CODIGO EN SISTEMA": 13, "DIAS": 6,
              "VALOR CARGO USD": 12, "SE VENDIO EN COUNTER": 30, "VALOR COUNTER COP": 14,
              "COMISION QUE CORRESPONDE": 15, "DIFERENCIA": 11, "QUIEN ENTREGO": 20,
              "LE PAGA HOY EL SISTEMA": 20, "LE CORRESPONDE": 12, "RESULTADO": 24,
              "QUE PASO": 90}
    money = ("VALOR", "VALOR COUNTER COP", "COMISION QUE CORRESPONDE", "DIFERENCIA")
    escribir_tabla(ws, rev, anchos, wrap_cols=("QUE PASO", "SE VENDIO EN COUNTER",
                                               "RESULTADO"),
                   money_cols=money, fila_ini=5)

    for i in range(6, 6 + len(rev)):
        res = ws.cell(row=i, column=list(rev.columns).index("RESULTADO") + 1).value
        fill = COLOR_RES.get(res)
        if fill:
            ws.cell(row=i, column=list(rev.columns).index("RESULTADO") + 1).fill = fill
        for j in range(1, 5):                       # columnas originales
            ws.cell(row=i, column=j).fill = ORIG
        ws.row_dimensions[i].height = 46
    ws.freeze_panes = "E6"
    ws.auto_filter.ref = "A5:Q" + str(5 + len(rev))

    # ---------- Hoja 2: resumen por asesor ----------
    res = []
    for a in ["STEFFANY", "DANILO", "NATALIA", "DAVID"]:
        s = rev[rev.ASESOR == a]
        num = pd.to_numeric(s["COMISION QUE CORRESPONDE"], errors="coerce").fillna(0)
        ok = s[s.RESULTADO.isin(["CORRECTO", "SUYO - FALTA DEFINIR OT/FI",
                                 "SUYO - REVISAR MONTO"])]
        okn = pd.to_numeric(ok["COMISION QUE CORRESPONDE"], errors="coerce").fillna(0)
        sin_otfi = ok[~ok.RESULTADO.str.contains("OT/FI")]
        res.append({
            "ASESOR": a,
            "LINEAS EN SU LISTA": len(s),
            "LINEAS VALIDAS": len(ok),
            "TOTAL QUE RECLAMO": TOTALES_DECLARADOS[a],
            "COMISION QUE LE CORRESPONDE": int(okn.sum()),
            "DE ESO, SIN CONTAR OT/FI": int(pd.to_numeric(
                sin_otfi["COMISION QUE CORRESPONDE"], errors="coerce").fillna(0).sum()),
            "DIFERENCIA VS LO QUE PIDIO": int(okn.sum()) - TOTALES_DECLARADOS[a],
        })
    resumen = pd.DataFrame(res)

    ws2 = wb.create_sheet("Resumen por asesor")
    ws2["A1"] = "RESUMEN POR ASESOR"
    ws2["A1"].font = Font(bold=True, size=14, color=INK, name="Calibri")
    ws2["A2"] = ("Solo sobre las lineas que ellos listaron. No incluye el resto de "
                 "contratos de julio que tambien estan mal atribuidos.")
    ws2["A2"].font = Font(size=10, italic=True, color="5A6472", name="Calibri")
    escribir_tabla(ws2, resumen,
                   {"ASESOR": 14, "LINEAS EN SU LISTA": 14, "LINEAS VALIDAS": 13,
                    "TOTAL QUE RECLAMO": 16, "COMISION QUE LE CORRESPONDE": 20,
                    "DE ESO, SIN CONTAR OT/FI": 18, "DIFERENCIA VS LO QUE PIDIO": 18},
                   money_cols=("TOTAL QUE RECLAMO", "COMISION QUE LE CORRESPONDE",
                               "DE ESO, SIN CONTAR OT/FI", "DIFERENCIA VS LO QUE PIDIO"),
                   fila_ini=4)

    julio = pd.DataFrame([
        ["Jeimmy Fajardo (Steffany)", 81805, 283224, 329312],
        ["Danilo Gutierrez",          46242, 107993, 199712],
        ["Natalia Quintero",          45774, 127497, 195374],
        ["David Bonilla",                 0,  42789,  63121],
    ], columns=["ASESOR", "COMISION QUE PAGA HOY EL SISTEMA",
                "COMISION CORREGIDA", "CORREGIDA SI OT Y FI COMISIONAN"])
    fila = 4 + len(resumen) + 3
    ws2.cell(row=fila - 1, column=1,
             value="JULIO COMPLETO (todos sus contratos, no solo los que reclamaron)"
             ).font = Font(bold=True, size=12, color=INK, name="Calibri")
    escribir_tabla(ws2, julio,
                   {"ASESOR": 26, "COMISION QUE PAGA HOY EL SISTEMA": 22,
                    "COMISION CORREGIDA": 20, "CORREGIDA SI OT Y FI COMISIONAN": 24},
                   money_cols=("COMISION QUE PAGA HOY EL SISTEMA", "COMISION CORREGIDA",
                               "CORREGIDA SI OT Y FI COMISIONAN"), fila_ini=fila)

    # ---------- Hoja 3: que hicimos mal ----------
    errores = pd.DataFrame([
        {"#": 1,
         "QUE PASO": "El sistema le paga la comision al asesor que RECIBE el carro en la "
                     "devolucion, no al que lo ENTREGA y vende los adicionales en el counter.",
         "POR QUE PASO": "Sixt entrega dos columnas de operador. En su modelo, 'checkout' "
                         "significa la SALIDA del vehiculo, o sea la entrega. Nosotros lo "
                         "leimos al reves y usamos la columna equivocada para liquidar.",
         "COMO LO COMPROBAMOS": "Sixt tiene otra tabla (la de vehiculos) donde las columnas "
                                "se llaman 'handover' y 'return' sin ambiguedad. Comparamos "
                                "2.467 contratos de 2026: coinciden todos, cero diferencias. "
                                "Ademas, la columna SISTEMA que llenaron los asesores calza "
                                "17 de 17 con el operador de devolucion.",
         "CUANTO PESA": "287 de 429 contratos de julio (67%) tienen personas distintas en "
                        "entrega y devolucion. De las lineas que reclamaron, 23 se le pagan "
                        "hoy a otra persona.",
         "COMO SE CORRIGE": "Leer el operador de entrega desde la tabla de vehiculos de Sixt "
                            "y renombrar las columnas para que no se vuelvan a confundir. "
                            "Luego reconstruir silver y reliquidar.",
         "ESTADO": "PENDIENTE"},
        {"#": 2,
         "QUE PASO": "En pesos, mas de la mitad de los adicionales que el cliente ya habia "
                     "pagado en la reserva aparecian como si se hubieran vendido en el counter.",
         "POR QUE PASO": "El calculo que separa 'prepagado' de 'counter' quedo bien en dolares "
                         "pero mal en pesos.",
         "COMO LO COMPROBAMOS": "766 de 1.461 lineas de cargos de julio tienen el valor "
                                "prepagado en dolares pero cero en pesos, con todo el monto "
                                "cargado a counter.",
         "CUANTO PESA": "Infla la base de comision de cualquier consulta hecha directo a la "
                        "base de datos. El tablero que ve el equipo tiene un parche y por eso "
                        "en pantalla se ve bien.",
         "COMO SE CORRIGE": "Bajar ese parche al proceso que construye los datos, para que "
                            "quede corregido en la fuente y no solo en la pantalla.",
         "ESTADO": "PENDIENTE"},
        {"#": 3,
         "QUE PASO": "Hay contratos con los mismos cargos cargados dos veces, lo que duplica "
                     "su valor.",
         "POR QUE PASO": "La llave con la que se cargan los cargos no alcanza a identificar "
                         "filas repetidas, y el proceso las inserta en vez de reemplazarlas.",
         "COMO LO COMPROBAMOS": "97 grupos repetidos, 260 filas de mas, en 34 contratos de "
                                "todo el historico.",
         "CUANTO PESA": "En julio solo cae un contrato (9523788459). Ninguno de los contratos "
                        "que reclamaron los asesores esta afectado.",
         "COMO SE CORRIGE": "Afinar la llave y recargar los cargos de esos 34 contratos.",
         "ESTADO": "PENDIENTE"},
        {"#": 4,
         "QUE PASO": "Los codigos OT y FI no generan comision en el sistema, pero los asesores "
                     "si los cuentan.",
         "POR QUE PASO": "No es un error tecnico: es una definicion de negocio que nunca se "
                         "cerro. La lista de codigos que comisionan no los incluye.",
         "COMO LO COMPROBAMOS": "7 de las lineas del Excel son OT o FI. Danilo anoto un FI "
                                "como si fuera 'UP'.",
         "CUANTO PESA": "Unos 608.000 pesos de bolsa adicional en julio para toda la sede.",
         "COMO SE CORRIGE": "Definir con el area comercial si comisionan. Sea cual sea la "
                            "respuesta, dejarla escrita en un solo lugar y comunicarla al equipo.",
         "ESTADO": "DECISION DE NEGOCIO"},
    ])
    ws3 = wb.create_sheet("Que hicimos mal")
    ws3["A1"] = "QUE ERRORES COMETIMOS Y COMO SE CORRIGEN"
    ws3["A1"].font = Font(bold=True, size=14, color=INK, name="Calibri")
    escribir_tabla(ws3, errores,
                   {"#": 4, "QUE PASO": 52, "POR QUE PASO": 52, "COMO LO COMPROBAMOS": 52,
                    "CUANTO PESA": 46, "COMO SE CORRIGE": 46, "ESTADO": 16},
                   wrap_cols=tuple(errores.columns), fila_ini=3)
    for i in range(4, 4 + len(errores)):
        ws3.row_dimensions[i].height = 118

    # ---------- Hoja 4: la columna SISTEMA ----------
    sis = []
    for a, lista in SISTEMA_OK.items():
        for cto in lista:
            if cto in ctx.index:
                c = ctx.loc[cto]
                sis.append({
                    "ASESOR": a, "CONTRATO": cto,
                    "LO PUSO COMO 'EL SISTEMA SI ME LO DIO'": "Si",
                    "QUIEN LO ENTREGO DE VERDAD": nm(c.entrego),
                    "QUIEN LO RECIBIO EN LA DEVOLUCION": nm(c.acredita_hoy),
                    "COINCIDE CON QUIEN RECIBIO": "Si" if int(c.acredita_hoy) == ASESOR_COD[a] else "No",
                })
    sisdf = pd.DataFrame(sis)
    ws4 = wb.create_sheet("La columna SISTEMA")
    ws4["A1"] = "QUE ERA REALMENTE LA COLUMNA 'SISTEMA' DEL EXCEL"
    ws4["A1"].font = Font(bold=True, size=14, color=INK, name="Calibri")
    ws4["A2"] = ("Los asesores listaron ahi los contratos que el sistema si les reconocia. "
                 "Al revisarlos, TODOS resultaron ser contratos donde ellos RECIBIERON el "
                 "carro en la devolucion, no donde lo entregaron. Esa lista fue la "
                 "confirmacion de que el error estaba en la columna de operador.")
    ws4["A2"].font = Font(size=10, italic=True, color="5A6472", name="Calibri")
    ws4["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws4.merge_cells("A2:F3")
    ws4.row_dimensions[2].height = 30
    escribir_tabla(ws4, sisdf,
                   {"ASESOR": 12, "CONTRATO": 14,
                    "LO PUSO COMO 'EL SISTEMA SI ME LO DIO'": 22,
                    "QUIEN LO ENTREGO DE VERDAD": 24,
                    "QUIEN LO RECIBIO EN LA DEVOLUCION": 26,
                    "COINCIDE CON QUIEN RECIBIO": 18},
                   wrap_cols=("LO PUSO COMO 'EL SISTEMA SI ME LO DIO'",), fila_ini=5)

    wb.save(DEST)
    print("Escrito:", DEST)
    print("Hojas:", wb.sheetnames)
    print()
    print(resumen.to_string(index=False))
    print()
    print("columna SISTEMA: coincide con quien recibio en",
          (sisdf["COINCIDE CON QUIEN RECIBIO"] == "Si").sum(), "de", len(sisdf))


if __name__ == "__main__":
    main()
